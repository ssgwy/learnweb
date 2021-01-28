'''
1.打开excel文档
2.读取excel文档的所有sheet
3.获得特定sheet的属性（如title,tabColor等）

打开excel文档命令load_workbook其他四个参数说明：
    filename, 
    read_only=False,    默认是false，如果为true，则打开的文档为只读文档，不可写
    keep_vba=KEEP_VBA,  是否保住VBA代码内容
    data_only=False,    是否保留单元格公式，如果为true，则在打开时把单元格转换为具体的值。
    keep_links=True     指定单元格外部连接是否保存
'''

import openpyxl

wb = openpyxl.load_workbook('first.xlsx')
print(type(wb))


print('获取所有工作表的名字')
print(wb.sheetnames,'\n')

for sheet in wb:
    print(sheet.title)

sheet = wb['倒2']
print(type(sheet))

print(sheet.title)
print(sheet.sheet_properties.tabColor)
#显示单元格C2的内容，
#在data_only=False模式下显示的是单元格公式
#在data_only=True模式下显示的是单元格公式值
print(sheet['C2'].value)

wb.close()