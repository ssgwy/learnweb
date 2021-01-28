'''
1.获取一个单元格的数据
2.设置一个单元格的数据
3.循环获取和设置多个单元格的数据
4.获取工作表中包含数据的范围（最大行和列）
'''

import openpyxl
wb = openpyxl.load_workbook('first.xlsx')

print(wb.sheetnames,'\n')
sheet = wb['加在第一位置']
c9 = sheet['C9']
print(type(c9))

#获取单元格的数据
print(f'C9:{c9.value}')
print(f'D10:{sheet["D10"].value}')
print(f'E11:{sheet["E11"].value}')

#添加数据
sheet['A1'].value = '你好'
sheet['B10'].value = '=sum(C3,D11)'

#第2种方法：
print(sheet.cell(row = 9,column = 3,value = 200).value)

for x in range(1,101):
    for y in range(1,101):
        cell = sheet.cell(row = x,column = y)
        if cell.value != None:
            print(f'cells({x},{y})',cell.value)
print('-----------------')
print(sheet.max_column)
print(sheet.max_row)

wb.save('first.xlsx')
wb.close()
