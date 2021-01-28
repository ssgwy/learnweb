'''
1.创建新的Workbook
2.获得默认的sheet
3.添加新的sheet
4.保存Workbook
'''

#引用
from openpyxl import Workbook
#在内存中创建一人Excel文档
wb = Workbook()
#获得默认的sheet，操作首个工作表
ws = wb.active
print(ws.title)     #title即工作表名称
ws.title = '我的标题'
#添加新的sheet
ws1 = wb.create_sheet('追加在最后')   #追加一个工作表
ws2 = wb.create_sheet('加在第一位置',0)     #在第1个位置插入工作表
ws3 = wb.create_sheet('倒2',-1)     #在倒数第2个位置插入工作表
#设置工作表名称背景颜色
ws.sheet_properties.tabColor = '1072BA'   #蓝色
ws3.sheet_properties.tabColor = 'FFDD00'   #黄色

wb.save('first.xlsx')
