'''
1.获取一定范围的表格（子表格）
2.输出表格中所有的cell
3.获取特定行和列的cell集合
4.获取多列和多行
'''

import openpyxl
wb = openpyxl.load_workbook('first.xlsx')
sheet = wb['加在第一位置']
print('获取子表格')
cell_range = sheet['J1':'K21']
print(type(cell_range))
print(cell_range)
print('输出表格中所有的cell')
for rows in cell_range:   #遍历行
    for cell in rows:    #遍历列
        print(cell.value)

print('获取特定行和列的cell集合')
#根据最大行确定结束范围
cols = sheet['J']
for col in cols: 
    print(col.value)
print('max_row:',sheet.max_row)

rows = sheet['21']
for row in rows:
    print(row.value)
print('max_col:',sheet.max_column)

print('获取多列和多行')
col2s = sheet['J:K']
print(col2s)
print(len(col2s))

row2s = sheet['20:21']
print(row2s)
print(len(row2s))
#先行后列
#sheet.iter_rows(min_row=1,max_col=3,max_row=2)
i = 1
for rows in sheet.iter_rows(min_row=1,max_col=3,max_row=2):
    for cell in rows:
        print(cell)
        cell.value = i
        i += 1

#先列后行
i = 1
for cols in sheet.iter_cols(min_row=3,max_col=3,max_row=4):
    for cell in cols:
        print(cell)
        cell.value = i
        i += 1

wb.save('first.xlsx')
wb.close()

