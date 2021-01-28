import openpyxl

wb = openpyxl.load_workbook('first.xlsx')

#创建工作表，如果title重复，会使用数字索引让title保证唯一
ws1 = wb.create_sheet('hello world')

#修改工作表
ws1.sheet_properties.tabColor = 'FF0000'
ws1.title = 'abcdefg'
print('所有工工作表：',wb.sheetnames)
index = wb.index(ws1)
print('index:',index)   #得到工作表的索引，从0开始

#复制工作表
wsc = wb.copy_worksheet(ws1)
wsc.title = '复制的工作表'
wsc.sheet_properties.tabColor = '0000BB'

#删除工作表
wb.remove(wsc)
del wb['倒2']


wb.save('first.xlsx')
wb.close()