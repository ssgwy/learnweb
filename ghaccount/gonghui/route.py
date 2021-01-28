from flask import render_template, redirect, url_for, request, abort, current_app, flash
from flask_login import login_user, current_user, logout_user, login_required
from gonghui.forms import LoginForm, RegisterForm, PasswdResetRequestForm,\
    PasswdResetForm, AccountAddForm, AccountQueryForm, CondolenceAddForm, CondolenceForm,\
        MemberForm, TestForm, MemberListForm, MemberInfoForm
from gonghui.models.user import User
from gonghui.models.condolence import Condolence
from gonghui.models.account import Account
from gonghui.models.member import Member
from gonghui.models.member_info import Member_info
from gonghui import db
from gonghui.email import send_email
from datetime import datetime
from sqlalchemy import func
import datetime
import time
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Border, Side, colors
import tkinter.filedialog
import os


@login_required
def index():
    return render_template('index.html', title='首页')


def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = LoginForm()
    # if form.validate_on_submit():
    if request.method == 'POST':
        if request.form['request_button'] == "登录":
            u = User.query.filter_by(username=form.username.data).first()
            if u is None or not u.check_password(form.password.data):
                flash('用户名或密码错误！')
                return redirect(url_for('login'))
            login_user(u, remember=form.remember_me.data)  # 此句让系统记住登录者状态
            next_page = request.args.get('next')  # next_page让系统记住使登录后返回之前浏览页面
            if next_page:
                return redirect(next_page)
            return redirect(url_for('index'))
        if request.form['request_button'] == "重置":
            return redirect(url_for('password_reset_request'))
    return render_template('login.html', title="欢迎使用工会财管系统，请先登录", form=form)


def logout():
    logout_user()
    flash("已登出")
    return redirect(url_for('login'))


@login_required
def manageruser():
    users = User.query.all()
    #form = RegisterForm()
    # if form.validate_on_submit():
    #    user = User(username=form.username.data, email=form.email.data)
    #    user.set_password(form.password.data)
    #    db.session.add(user)
    #    db.session.commit()
    #    return redirect(url_for('manageruser'))
    return render_template('manageruser.html', title="用户管理", users=users)


def register():
    form = RegisterForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, email=form.email.data)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        return redirect(url_for('manageruser'))
    return render_template('register.html', title='增加用户', form=form)


def password_reset_request():
    # if current_user.is_authonticated:
    #    return redirect(url_for('index'))
    form = PasswdResetRequestForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user:
            flash("您将收到一封允许您重置密码的电子邮件。\
                如果你找不到这封邮件，请务必检查您的垃圾邮件！")
            token = user.get_jwt()
            url_password_reset = url_for(
                'password_reset',
                token=token,
                _external=True  # 发送的链接是完整的http链接
            )
            url_password_reset_request = url_for(
                'password_reset_request',
                _external=True
            )
            send_email(
               subject=current_app.config['MAIL_SUBJECT_PASSWORD_RESET'],  # 主题
               recipients=[user.email],  # 接收者
               text_body=render_template(
                   'email/passwd_reset.txt',
                   url_password_reset=url_password_reset,
                   url_password_reset_request=url_password_reset_request
               ),
               html_body=render_template(
                   'email/passwd_reset.html',
                   url_password_reset=url_password_reset,
                   url_password_reset_request=url_password_reset_request
               )
            )
        else:
            raise
        return redirect(url_for('login'))
    return render_template('password_reset_request.html', title='请求重置密码', form=form)


def password_reset(token):
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    user = User.verify_jwt(token)
    if not user:
        return redirect(url_for('login'))
    form = PasswdResetForm()
    if form.validate_on_submit():
        user.set_password(form.password.data)
        db.session.commit()
        flash("密码已重置")
        return redirect(url_for('login'))
    return render_template('password_reset.html', title='密码重置', form=form)


@login_required
def account_add():
    form = AccountAddForm()
    if request.method == 'POST':
        if float(form.a_type.data[:3]) < 1:
            in_out = "收入"
        else:
            in_out = "支出"
        account = Account(a_date=form.a_date.data, matters=form.matters.data, a_type=form.a_type.data,
            a_count=form.a_count.data, in_out=in_out, remark=form.remark.data)
        db.session.add(account)
        db.session.commit()
        flash(str(form.a_date.data) + form.matters.data + "已增加")
        return redirect(url_for('account_add'))
    return render_template('account_add.html', title="财务记账", form=form)


@login_required
def account_query():
    form = AccountQueryForm()
    if request.method == 'POST':
        out_kaiguan = 0
        filters = []
        if form.date_begin.data and form.date_end.data:
            filters.append(Account.a_date.between(
                form.date_begin.data, form.date_end.data))
        else:
            if form.date_begin.data:
                filters.append(Account.a_date >= form.date_begin.data)
            if form.date_end.data:
                filters.append(Account.a_date <= form.date_end.data)
        if form.matters.data:
            filters.append(Account.matters.like(
                "%{}%".format(form.matters.data)))
        if form.a_type.data:
            filters.append(Account.a_type == form.a_type.data)
        if form.in_out.data:
            filters.append(Account.in_out == form.in_out.data)
        accounts = Account.query.order_by(
            Account.a_date).filter(*filters).all()
        if accounts is None:
            abort(404)
        # 统计时间段内收入支出合计
        filters_in = []
        filters_out = []
        filters_in = filters[:]
        filters_out = filters[:]
        filters_in.append(Account.in_out == '收入')
        filters_out.append(Account.in_out == '支出')
        sum_in = db.session.query(
            func.sum(Account.a_count)).filter(*filters_in).all()
        sum_out = db.session.query(
            func.sum(Account.a_count)).filter(*filters_out).all()
        count_in = db.session.query(func.count(
            Account.a_count)).filter(*filters_in).all()
        count_out = db.session.query(func.count(
            Account.a_count)).filter(*filters_out).all()
        if sum_in[0][0] is None:
            sum_in_value = 0
            num_count_in = 0
        else:
            sum_in_value = sum_in[0][0]
            num_count_in = count_in[0][0]
        if sum_out[0][0] is None:
            sum_out_value = 0
            num_count_out = 0
        else:
            sum_out_value = sum_out[0][0]
            num_count_out = count_out[0][0]
        all_count = sum_in_value - sum_out_value
        num_count_all = num_count_in + num_count_out
    # ---------------------------以下功能区-------------------------------------
        if request.form['request_button'] == "查询":
            #    i = 1
            #    for account in accounts:
            #        account.id = i
            #        i +=1
            return render_template('account_query.html', title="账务查询", accounts=accounts, form=form, sum_in_value=sum_in_value, sum_out_value=sum_out_value, all_count=all_count, num_count_in=num_count_in, num_count_out=num_count_out, num_count_all=num_count_all, is_ok=0)
        if request.form['request_button'] == "导出Excel":
            if len(filters) == 0:
                flash("请输入查询条件")
                return redirect(url_for('account_query'))
            if accounts is None:
                flash('查无数据！')
                return redirect(url_for('account_query'))
            else:
                wb = Workbook()
                sheet = wb.active
                sheet.title = '记账查询'
                sheet['A1'].value = '序号'
                sheet['B1'].value = '记账日期'
                sheet['C1'].value = '记账事项'
                sheet['D1'].value = '会计科目'
                sheet['E1'].value = '记账金额'
                sheet['F1'].value = '备注'
                for accout in accounts:
                    # print(con.id,con.c_date,con.c_name,con.c_type,con.c_count,con.remark,con.in_account)
                    i = sheet.max_row + 1
                    sheet.cell(row=i, column=1).value = i-1
                    sheet.cell(row=i, column=2).value = accout.a_date
                    sheet.cell(row=i, column=3).value = accout.matters
                    sheet.cell(row=i, column=4).value = accout.a_type
                    sheet.cell(row=i, column=5).value = accout.a_count
                    sheet.cell(row=i, column=6).value = accout.remark
                #file_path = tkinter.filedialog.asksaveasfilename(filetypes=[('xlsx', '*.xlsx')],initialdir='D:\\')
                #filename = filename + '.xls'
                sheet.column_dimensions['A'].width = 4.5
                sheet.column_dimensions['B'].width = 12
                sheet.column_dimensions['C'].width = 34
                sheet.column_dimensions['D'].width = 23
                sheet.column_dimensions['E'].width = 10
                sheet.column_dimensions['F'].width = 30
                wb.save("D:\\记账数据导出" +
                        str(time.strftime('%Y-%m-%d %H.%M.%S')) + '.xlsx')
                wb.close()
                flash('已导出到' + "D:\\记账数据" +
                      str(time.strftime('%Y-%m-%d %H.%M.%S')) + '.xlsx')
            return render_template('account_query.html', title="账务查询", accounts=accounts, form=form, sum_in_value=sum_in_value, sum_out_value=sum_out_value, all_count=all_count, num_count_in=num_count_in, num_count_out=num_count_out, num_count_all=num_count_all, is_ok=0)
        if request.form['request_button'] == "删除":
            acco = Account.query.order_by(
                Account.id.desc()).filter(*filters).first()
            f_acco = []
            f_acco.append(acco)
            return render_template('account_query.html', title="账务查询", is_ok=1, accounts=f_acco, form=form)
        if request.form['request_button'] == "确认":
            if form.input_ok.data == 'OK':
                accounts = Account.query.order_by(
                    Account.id.desc()).filter(*filters).first()
                db.session.delete(accounts)
                if accounts.matters.find('季度慰问') > 0:
                    y_end = accounts.a_date.strftime('%Y')
                    m_end = accounts.a_date.strftime('%m')
                    if int(m_end) == 12:
                        m_start = '10'
                    else:
                        m_start = '0' + str(int(m_end)-2)
                    date_start = datetime.datetime.strptime(
                        y_end + "-" + m_start + "-01", "%Y-%m-%d").strftime("%Y-%m-%d")
                    c = db.session.query(Condolence.in_account).filter(Condolence.c_date.between(
                        date_start, accounts.a_date)).update({'in_account': ''}, synchronize_session=False)
                flash(str(accounts.a_date) + accounts.matters + "已删除")
                db.session.commit()
                return render_template('account_query.html', title="账务查询", is_ok=0, form=form)
            else:
                acco = Account.query.order_by(
                    Account.id.desc()).filter(*filters).first()
                f_acco = []
                f_acco.append(acco)
                flash("出错：请输入大写的OK，再点击【确认】")
                return render_template('account_query.html', title="账务查询", is_ok=1, accounts=f_acco, form=form)
        if request.form['request_button'] == "取消":
            return render_template('account_query.html', title="账务查询", is_ok=0, form=form)
        if request.form['request_button'] == "导入记账信息":
            if os.access("D:/工会财管/记账数据导出.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('D:/工会财管/记账数据导出.xlsx')
                sheet = wb['记账查询']
                if sheet.max_row == 1:
                    flash("数据表为空表！")
                    return redirect(url_for('account_query'))
                for x in range(2, sheet.max_row+1):
                    id = sheet.cell(row=x, column=1).value
                    a_date = sheet.cell(
                        row=x, column=2).value.strftime('%Y-%m-%d')
                    matters = sheet.cell(row=x, column=3).value
                    a_type = sheet.cell(row=x, column=4).value
                    a_count = sheet.cell(row=x, column=5).value
                    if float(a_type[:3]) < 1:
                        in_out = "收入"
                    else:
                        in_out = "支出"
                    remark = sheet.cell(row=x, column=6).value
                    if remark is None:
                        remark = ''
                    else:
                        remark = remark
                    account = Account(a_date=a_date, matters=matters, a_type=a_type,
                                      a_count=a_count, in_out=in_out, remark=remark)
                    db.session.add(account)
                db.session.commit()
                #wb.save('D:/工会财管/记账数据导出（已导入' + str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + '）.xlsx')
                wb.close()
                os.rename("D:/工会财管/记账数据导出.xlsx", "D:/工会财管/记账数据导入" +
                          str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + ".xlsx")
                flash("记账信息批量导入完成")
            else:
                flash("文档不存在！")
            return render_template('account_query.html', title="账务查询", form=form, is_ok=0)
    return render_template('account_query.html', title="账务查询", form=form, is_ok=0)


@login_required
def condolence_add():
    form = CondolenceAddForm()
    if request.method == 'POST':
        # if form.validate_on_submit():
        condolence = Condolence(c_date=form.c_date.data, c_year=form.c_date.data.strftime('%Y'), c_quarter=len(str(2**int(form.c_date.data.strftime('%m')))), c_name=form.c_name.data,
            c_type=form.c_type.data, c_count=form.c_count.data, remark=form.remark.data, in_account="未入账")
        db.session.add(condolence)
        db.session.commit()
        flash(str(form.c_date.data) + form.c_name.data +
              form.c_type.data + str(int(form.c_count.data)) + "元 已增加")
        return redirect(url_for('condolence_add'))
    return render_template('condolence_add.html', title="慰问登记", form=form)


@login_required
def condolence():
    form = CondolenceForm()
    if request.method == 'POST':
        # if form.validate_on_submit():
        filters = []
        str_filter = '慰问导出数据('
        if form.date_begin.data and form.date_end.data:
            filters.append(Condolence.c_date.between(
                form.date_begin.data, form.date_end.data))
        else:
            if form.date_begin.data:
                filters.append(Condolence.c_date >= form.date_begin.data)
            if form.date_end.data:
                filters.append(Condolence.c_date <= form.date_end.data)
        if form.c_name.data:
            filters.append(Condolence.c_name.like(
                "%{}%".format(form.c_name.data)))
        if form.c_type.data:
            filters.append(Condolence.c_type == form.c_type.data)
        if form.in_account.data:
            filters.append(Condolence.in_account == form.in_account.data)
        condolences = Condolence.query.order_by(
            Condolence.c_date).filter(*filters).all()
        condolence_counts = db.session.query(
            func.count(Condolence.c_date)).filter(*filters).all()
        all_count = db.session.query(
            func.sum(Condolence.c_count)).filter(*filters).all()
        if all_count[0][0] is None:
            all_count_value = 0
            condolence_count = 0
        else:
            all_count_value = all_count[0][0]
            condolence_count = condolence_counts[0][0]
        # -------------------------------------------各个按钮的处理
        if request.form['request_button'] == "查询":
            return render_template('condolence.html', title="慰问查询", all_count_value=all_count_value, condolence_count=condolence_count, condolences=condolences, form=form, is_ok=0)
        if request.form['request_button'] == "导出Excel":
            if len(filters) == 0:
                flash("请输入查询条件！")
                return redirect(url_for('condolence'))
            if condolences is None:
                flash('查无数据！')
                return redirect(url_for('condolence'))
            else:
                wb = Workbook()
                sheet = wb.active
                sheet.title = '慰问查询'
                sheet['A1'].value = '序号'
                sheet['B1'].value = '慰问日期'
                sheet['C1'].value = '被慰问人'
                sheet['D1'].value = '慰问类型'
                sheet['E1'].value = '慰问金额'
                sheet['F1'].value = '慰问备注'
                sheet['G1'].value = '是否入账'
                for con in condolences:
                    # print(con.id,con.c_date,con.c_name,con.c_type,con.c_count,con.remark,con.in_account)
                    i = sheet.max_row + 1
                    sheet.cell(row=i, column=1).value = i-1
                    sheet.cell(row=i, column=2).value = con.c_date
                    sheet.cell(row=i, column=3).value = con.c_name
                    sheet.cell(row=i, column=4).value = con.c_type
                    sheet.cell(row=i, column=5).value = con.c_count
                    sheet.cell(row=i, column=6).value = con.remark
                    sheet.cell(row=i, column=7).value = con.in_account
                #file_path = tkinter.filedialog.asksaveasfilename(filetypes=[('xlsx', '*.xlsx')],initialdir='D:\\')
                #filename = filename + '.xls'
                sheet.column_dimensions['A'].width = 4.5
                sheet.column_dimensions['B'].width = 12
                sheet.column_dimensions['C'].width = 9
                sheet.column_dimensions['D'].width = 9
                sheet.column_dimensions['F'].width = 38
                wb.save("D:/工会财管/慰问数据导出" +
                        str(time.strftime('%Y-%m-%d %H.%M.%S')) + '.xlsx')
                wb.close()
                flash('已导出到' + "D:/工会财管/慰问数据" +
                      str(time.strftime('%Y-%m-%d %H.%M.%S')) + '.xlsx')
            return render_template('condolence.html', title="慰问查询", all_count_value=all_count_value, condolence_count=condolence_count, condolences=condolences, form=form, is_ok=0)
        if request.form['request_button'] == "季结入账":
            # 填写格式检查，并根据填写起止日期确定所属季度，当季最后一天为季结入账时间
            if form.date_begin.data == None or form.date_end.data == None:
                flash("请选择起止日期，且必须为当季第一天和最后一天")
                return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)
            else:
                dt_begin_year = form.date_begin.data.strftime('%Y')
                dt_begin_month = form.date_begin.data.strftime('%m')
                dt_begin_day = form.date_begin.data.strftime('%d')
                dt_end_year = form.date_end.data.strftime('%Y')
                dt_end_month = form.date_end.data.strftime('%m')
                dt_end_day = form.date_end.data.strftime('%d')
                if int(dt_end_year) != int(dt_begin_year):
                    flash("操作失败：开始和结束日期必须为同一年份")
                    return render_template('condolence.html', title="慰问查询", form=form)
                if int(dt_begin_month) % 3 != 1:
                    flash("操作失败：开始日期" + dt_begin_month + "月不是季度首月")
                    return render_template('condolence.html', title="慰问查询", form=form)
                if int(dt_begin_day) != 1:
                    flash("操作失败：开始日期" + dt_begin_day + "日不是季度首月1日")
                    return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)
                if int(dt_end_month) % 3 != 0:
                    flash("操作失败：结束日期" + dt_end_month + "月不是季度末月")
                    return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)
                if (int(dt_end_month) == 3 or int(dt_end_month) == 12) and int(dt_end_day) != 31:
                    flash("操作失败：结束日期" + dt_end_month + "月最后一天必须为31日")
                    return render_template('condolence.html', title="慰问查询", form=form)
                if (int(dt_end_month) == 6 or int(dt_end_month) == 9) and int(dt_end_day) != 30:
                    flash("操作失败：结束日期" + dt_end_month + "月最后一天必须为30日")
                    return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)
                if int(dt_end_month) - int(dt_begin_month) != 2:
                    flash("操作失败：开始和结束日期必须为同一季度内月份")
                    return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)
                if form.c_name.data or form.c_type.data:
                    form.c_name.data = ''
                    form.c_type.data = ''
                    flash("操作失败：季结入账，不必选择慰问人和慰问类型")
                    return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)
                if all_count_value == 0:
                    flash("操作失败：季结慰问为0，无需入账")
                    return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)
                # 当季不存在未入账记录，无须入账
                e_filters = []
                e_filters.append(Condolence.in_account != "已入账")
                condo = Condolence.query.filter(*e_filters).first()
                if condo == None:
                    flash("操作失败：当前季度慰问已全部入账，无须二次入账")
                    return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)
                matters = str(dt_end_year) + "年第" + \
                              str(len(str(2**int(dt_end_month)))) + "季度慰问"
                e = Account.query.filter_by(matters=matters).first()
                if e:
                    # 如果已经季结，则更新季结金额，并把当条记录in_account更新为“已记账”
                    e.a_count = all_count_value
                    db.session.commit()
                    c = db.session.query(Condolence.in_account).filter(
                        *filters).update({'in_account': '已入账'}, synchronize_session=False)
                    db.session.commit()
                    flash(matters + " 已更新")
                else:
                    account = Account(a_date=form.date_end.data, matters=matters,
                                      a_type="1.4 职工集体福利支出", a_count=all_count_value, in_out="支出", remark="慰问")
                    db.session.add(account)
                    db.session.commit()
                    c = db.session.query(Condolence.in_account).filter(
                        *filters).update({'in_account': '已入账'}, synchronize_session=False)
                    db.session.commit()
                    flash(matters + " 已增加")
                return render_template('condolence.html', title="慰问查询", all_count_value=all_count_value, condolences=condolences, form=form, is_ok=0)
        if request.form['request_button'] == "删除":
            condo = Condolence.query.order_by(
                Condolence.id.desc()).filter(*filters).first()
            if condo != None:
                f_condo = []
                f_condo.append(condo)
                return render_template('condolence.html', title="慰问查询", is_ok=1, condolences=f_condo, form=form)
            flash("未找到相关记录，无法进行删除操作！")
            return render_template('condolence.html', title="慰问查询", is_ok=0, form=form)
        if request.form['request_button'] == "确认":
            if form.is_ok.data == 'OK':
                condo = Condolence.query.order_by(
                    Condolence.id.desc()).filter(*filters).first()
                condo_year = condo.c_year
                condo_quarter = condo.c_quarter
                condo_account = condo.c_count
                matters = str(condo_year) + "年第" + str(condo_quarter) + "季度慰问"
                db.session.delete(condo)
                flash(str(condo.c_date) + condo.c_name + condo.c_type + " 已删除")
                db.session.commit()
                if condo.in_account == "已入账":
                    ec = Account.query.filter_by(matters=matters).first()
                    ec.a_count = ec.a_count - float(condo_account)
                    db.session.commit()
                condolences = Condolence.query.order_by(
                    Condolence.c_date).filter(*filters).all()
                all_count = db.session.query(
                    func.sum(Condolence.c_count)).filter(*filters).all()
                if all_count[0][0] is None:
                    all_count_value = 0
                else:
                    all_count_value = all_count[0][0]
                return render_template('condolence.html', title="慰问查询", is_ok=0, all_count_value=all_count_value, condolences=condolences, form=form)
        if request.form['request_button'] == "取消":
            return render_template('condolence.html', title="慰问查询", is_ok=0, form=form)
        if request.form['request_button'] == "导入Excel":
            if os.access("D:/工会财管/慰问数据导出.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('慰问数据导出.xlsx')
                sheet = wb['慰问查询']
                if sheet.max_row == 1:
                    flash("数据表为空表！")
                    return redirect(url_for('condolence'))
                for x in range(2, sheet.max_row+1):
                    id = sheet.cell(row=x, column=1).value
                    c_date = sheet.cell(
                        row=x, column=2).value.strftime('%Y-%m-%d')
                    c_year = sheet.cell(row=x, column=2).value.strftime('%Y')
                    c_quarter = len(
                        str(2**int(sheet.cell(row=x, column=2).value.strftime('%m'))))
                    c_name = sheet.cell(row=x, column=3).value
                    c_type = sheet.cell(row=x, column=4).value
                    c_count = sheet.cell(row=x, column=5).value
                    remark = sheet.cell(row=x, column=6).value
                    if remark is None:
                        remark = ''
                    else:
                        remark = remark
                    in_account = sheet.cell(row=x, column=7).value
                    condolence = Condolence(c_date=c_date, c_year=c_year, c_quarter=c_quarter, c_name=c_name,
                        c_type=c_type, c_count=c_count, remark=remark, in_account=in_account)
                    db.session.add(condolence)
                db.session.commit()
                wb.close()
                os.rename("D:/工会财管/慰问数据导出.xlsx", "D:/工会财管/慰问数据导入" +
                          str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + ".xlsx")
                flash("慰问信息批量导入完成")
            else:
                flash("文档不存在！")
            return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)
    return render_template('condolence.html', title="慰问查询", form=form, is_ok=0)


@login_required
def test():
    form = TestForm()
    date_year = form.date_year.data
    date_half_year = form.date_half_year.data
    if request.method == 'POST':
        if request.form['request_button'] == "导入记账信息":
            if os.access("D:/工会财管/记账数据导出.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('D:/工会财管/记账数据导出.xlsx')
                sheet = wb['记账查询']
                if sheet.max_row == 1:
                    flash("数据表为空表！")
                    return redirect(url_for('test'))
                for x in range(2, sheet.max_row+1):
                    id = sheet.cell(row=x, column=1).value
                    a_date = sheet.cell(
                        row=x, column=2).value.strftime('%Y-%m-%d')
                    matters = sheet.cell(row=x, column=3).value
                    a_type = sheet.cell(row=x, column=4).value
                    a_count = sheet.cell(row=x, column=5).value
                    if float(a_type[:3]) < 1:
                        in_out = "收入"
                    else:
                        in_out = "支出"
                    remark = sheet.cell(row=x, column=6).value
                    if remark is None:
                        remark = ''
                    else:
                        remark = remark
                    account = Account(a_date=a_date, matters=matters, a_type=a_type,
                                      a_count=a_count, in_out=in_out, remark=remark)
                    db.session.add(account)
                db.session.commit()
                #wb.save('D:/工会财管/记账数据导出（已导入' + str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + '）.xlsx')
                wb.close()
                os.rename("D:/工会财管/记账数据导出.xlsx", "D:/工会财管/记账数据导入" +
                          str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + ".xlsx")
                flash("记账信息批量导入完成")
            else:
                flash("文档不存在！")
            return render_template('test.html', title='测试网页代码', form=form)
        if request.form['request_button'] == "导入Excel":
            if os.access("D:/工会财管/慰问数据导出.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('慰问数据导出.xlsx')
                sheet = wb['慰问查询']
                if sheet.max_row == 1:
                    flash("数据表为空表！")
                    return redirect(url_for('test'))
                for x in range(2, sheet.max_row+1):
                    id = sheet.cell(row=x, column=1).value
                    c_date = sheet.cell(
                        row=x, column=2).value.strftime('%Y-%m-%d')
                    c_year = sheet.cell(row=x, column=2).value.strftime('%Y')
                    c_quarter = len(
                        str(2**int(sheet.cell(row=x, column=2).value.strftime('%m'))))
                    c_name = sheet.cell(row=x, column=3).value
                    c_type = sheet.cell(row=x, column=4).value
                    c_count = sheet.cell(row=x, column=5).value
                    remark = sheet.cell(row=x, column=6).value
                    if remark is None:
                        remark = ''
                    else:
                        remark = remark
                    in_account = sheet.cell(row=x, column=7).value
                    condolence = Condolence(c_date=c_date, c_year=c_year, c_quarter=c_quarter, c_name=c_name,
                        c_type=c_type, c_count=c_count, remark=remark, in_account=in_account)
                    db.session.add(condolence)
                db.session.commit()
                wb.close()
                os.rename("D:/工会财管/慰问数据导出.xlsx", "D:/工会财管/慰问数据导入" +
                          str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + ".xlsx")
                flash("慰问信息批量导入完成")
            else:
                flash("文档不存在！")
            return render_template('test.html', title='测试网页代码', form=form)
        if request.form['request_button'] == "导入会费信息":
            if os.access("D:/工会财管/会费信息.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('D:/工会财管/会费信息.xlsx')
                sheet = wb['会费信息']
                if sheet.max_row == 1:
                    flash("数据表为空表！")
                    return redirect(url_for('test'))
                for x in range(2, sheet.max_row+1):
                    member_id = sheet.cell(row=x, column=1).value
                    date_year = sheet.cell(
                        row=2, column=7).value  # .strftime('%Y')
                    date_half_year = sheet.cell(row=2, column=8).value
                    name = sheet.cell(row=x, column=2).value
                    ID_num = str(sheet.cell(row=x, column=4).value)
                    salary = sheet.cell(row=x, column=5).value
                    membership_fee = sheet.cell(
                        row=x, column=5).value * 0.005 * 6
                    personnel_type = sheet.cell(row=x, column=3).value
                    if int(membership_fee) == membership_fee:
                        membership_fee = membership_fee
                    else:
                        membership_fee = int(membership_fee) + 1
                    # print(id,date_year,date_half_year,name,ID_num,salary,membership_fee,personnel_type)
                    member = Member(member_id=member_id, date_year=date_year, date_half_year=date_half_year, name=name,
                                    ID_num=ID_num, salary=salary, membership_fee=membership_fee, personnel_type=personnel_type)
                    db.session.add(member)
                db.session.commit()
                wb.close()
                os.rename("D:/工会财管/会费信息.xlsx", "D:/工会财管/会费信息导入" +
                          str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + ".xlsx")
                flash("会费信息批量导入完成")
            else:
                flash("会费文档不存在！")
            return render_template('test.html', title='测试网页代码', form=form)
        if request.form['request_button'] == "导入会员信息":
            if os.access("D:/工会财管/会员信息.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('D:/工会财管/会员信息.xlsx')
                sheet = wb['会员信息']
                if sheet.max_row == 1:
                    flash("数据表为空表！")
                    return redirect(url_for('test'))
                for x in range(2, sheet.max_row+1):
                    id = sheet.cell(row=x, column=1).value
                    # .strftime('%Y')
                    ID_num = str(sheet.cell(row=x, column=3).value)
                    if sheet.cell(row=x, column=4).value is None:
                        date_birth = ''
                    else:
                        date_birth = sheet.cell(row=x, column=4).value
                    if sheet.cell(row=x, column=5).value is None:
                        gender = ''
                    else:
                        gender = sheet.cell(row=x, column=5).value
                    if sheet.cell(row=x, column=6).value is None:
                        nation = ''
                    else:
                        nation = sheet.cell(row=x, column=6).value
                    if sheet.cell(row=x, column=7).value is None:
                        political_status = ''
                    else:
                        political_status = sheet.cell(row=x, column=7).value
                    if sheet.cell(row=x, column=8).value is None:
                        native_palce = ''
                    else:
                        native_place = sheet.cell(row=x, column=8).value
                    if sheet.cell(row=x, column=9).value is None:
                        education = ''
                    else:
                        education = sheet.cell(row=x, column=9).value
                    if sheet.cell(row=x, column=10).value is None:
                        phone = ''
                    else:
                        phone = sheet.cell(row=x, column=10).value
                    if sheet.cell(row=x, column=11).value is None:
                        email = ''
                    else:
                        email = sheet.cell(row=x, column=11).value
                    if sheet.cell(row=x, column=12).value is None:
                        remark = ''
                    else:
                        remark = sheet.cell(row=x, column=12).value
                    # print(id,ID_num,date_birth,gender,nation,political_status,native_place,education,phone,email,remark)
                    member_info = Member_info(ID_num=ID_num, date_birth=date_birth, gender=gender, nation=nation, political_status=political_status,
                                              native_place=native_place, education=education, phone=phone, email=email, remark=remark)
                    db.session.add(member_info)
                db.session.commit()
                wb.close()
                os.rename("D:/工会财管/会员信息.xlsx", "D:/工会财管/会员信息导入" +
                          str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + ".xlsx")
                flash("会员信息批量导入完成")
            else:
                flash("会员文档不存在！")
            return render_template('test.html', title='测试网页代码', form=form)
        if request.form['request_button'] == "导出互助数据":
            filters = []
            if form.date_year.data:
                filters.append(Member.date_year == form.date_year.data)
            else:
                flash('请输入年份')
                return render_template('test.html', title='测试网页代码', form=form)
            if form.date_half_year.data:
                filters.append(Member.date_half_year ==
                               form.date_half_year.data)
            else:
                flash('请输入半年区间')
                return render_template('test.html', title='测试网页代码', form=form)
            members = db.session.query(Member.id, Member.name, Member_info.gender, Member.ID_num, Member_info.phone).join(
                Member_info, Member.ID_num == Member_info.ID_num).filter(*filters).all()
            #members = Member_info.query.order_by(Member_info.id).filter(*filters).all()
            if os.access("D:/工会财管/数据模板/机关事业-模板.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('D:/工会财管/数据模板/机关事业-模板.xlsx')
                sheet = wb['事业2']
                i = 3
                for member in members:
                    i = i + 1
                    print(i-3, member.name, member.gender,
                          member.ID_num, member.phone)
                    sheet.cell(row=i, column=1, value=i-3)
                    sheet.cell(row=i, column=2, value=member.name)
                    sheet.cell(row=i, column=3, value=member.gender)
                    sheet.cell(row=i, column=4, value=member.ID_num)
                    sheet.cell(row=i, column=5, value=member.phone)
                wb.save('D:/工会财管/石狮市永宁中学-机关事业-1607395530.xlsx')
                wb.close()
                flash("互助信息已导出，文件名：D:/工会财管/数据模板/石狮市永宁中学-机关事业-1607395530.xlsx")
            else:
                flash("互助模板不存在，文件名：机关事业-模板.xlsx")
            return render_template('test.html', title='测试网页代码', form=form)
    return render_template('test.html', title='测试网页代码', form=form)


@login_required
def member_manager():
    form = MemberForm()
    if request.method == 'POST':
        if request.form['request_button'] == "批量导入会费信息":
            wb = openpyxl.load_workbook('D:/工会财管/会费信息.xlsx')
            sheet = wb['会费信息']
            for x in range(2, sheet.max_row+1):
                id = sheet.cell(row=x, column=1).value
                a_date = sheet.cell(row=x, column=2).value.strftime('%Y-%m-%d')
                matters = sheet.cell(row=x, column=3).value
                a_type = sheet.cell(row=x, column=4).value
                a_count = sheet.cell(row=x, column=5).value
                if float(a_type[:3]) < 1:
                    in_out = "收入"
                else:
                    in_out = "支出"
                remark = sheet.cell(row=x, column=6).value
                if remark is None:
                    remark = ''
                else:
                    remark = remark
                print(id, a_date, matters, a_type, a_count, in_out, remark)
                print(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime()))
                '''
                account = Account(a_date=a_date,matters=matters,a_type=a_type,
                                  a_count=a_count,in_out=in_out,remark=remark)
                db.session.add(account)
            db.session.commit()
            '''
            #wb.save('D:/工会财管/记账数据导出（已导入' + str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + '）.xlsx')
            wb.close()
            os.rename("D:/工会财管/记账数据导出.xlsx", "D:/工会财管/记账数据导出（已导入" +
                      str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + "）.xlsx")
            flash("记账信息批量导入完成")

        if request.form['request_button'] == "批量导入会员信息":
            wb = openpyxl.load_workbook('D:/工会财管/会员信息.xlsx')
            sheet = wb['会员信息']
            for x in range(2, sheet.max_row+1):
                id = sheet.cell(row=x, column=1).value
                a_date = sheet.cell(row=x, column=2).value.strftime('%Y-%m-%d')
                matters = sheet.cell(row=x, column=3).value
                a_type = sheet.cell(row=x, column=4).value
                a_count = sheet.cell(row=x, column=5).value
                if float(a_type[:3]) < 1:
                    in_out = "收入"
                else:
                    in_out = "支出"
                remark = sheet.cell(row=x, column=6).value
                if remark is None:
                    remark = ''
                else:
                    remark = remark
                print(id, a_date, matters, a_type, a_count, in_out, remark)
                print(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime()))
                '''
                account = Account(a_date=a_date,matters=matters,a_type=a_type,
                                  a_count=a_count,in_out=in_out,remark=remark)
                db.session.add(account)
            db.session.commit()
            '''
            #wb.save('D:/工会财管/记账数据导出（已导入' + str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + '）.xlsx')
            wb.close()
            os.rename("D:/工会财管/记账数据导出.xlsx", "D:/工会财管/记账数据导出（已导入" +
                      str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + "）.xlsx")
            flash("记账信息批量导入完成")

            return render_template('test.html', title='测试网页代码', form=form)
    return render_template('member_manager.html', title="会员管理", form=form)


@login_required
def membership_list():
    form = MemberListForm()
    if request.method == 'POST':
        filters = []
        if form.date_year.data:
            filters.append(Member.date_year == form.date_year.data)
        else:
            flash('请输入年份')
            return render_template('membership_list.html', title='会费情况', form=form, is_ok=0)
        if form.date_half_year.data:
            filters.append(Member.date_half_year == form.date_half_year.data)
        else:
            flash('请输入半年区间')
            return render_template('membership_list.html', title='会费情况', form=form, is_ok=0)
        if form.name.data:
            filters.append(Member.name == form.name.data)
        if form.personnel_type.data:
            filters.append(Member.personnel_type == form.personnel_type.data)
        members = Member.query.order_by(
            Member.member_id).filter(*filters).all()
        if request.form['request_button'] == "查询":
            return render_template('membership_list.html', title='会费情况', form=form, members=members, year=form.date_year.data, half_year=form.date_half_year.data, is_ok=0)
        if request.form['request_button'] == "导入Excel":
            if os.access("D:/工会财管/会费信息.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('D:/工会财管/会费信息.xlsx')
                sheet = wb['会费信息']
                if sheet.max_row == 1:
                    flash("数据表为空表！")
                    return redirect(url_for('membership_list'))
                for x in range(2, sheet.max_row+1):
                    member_id = sheet.cell(row=x, column=1).value
                    date_year = sheet.cell(
                        row=2, column=7).value  # .strftime('%Y')
                    date_half_year = sheet.cell(row=2, column=8).value
                    name = sheet.cell(row=x, column=2).value
                    ID_num = str(sheet.cell(row=x, column=4).value)
                    salary = sheet.cell(row=x, column=5).value
                    membership_fee = sheet.cell(
                        row=x, column=5).value * 0.005 * 6
                    personnel_type = sheet.cell(row=x, column=3).value
                    if int(membership_fee) == membership_fee:
                        membership_fee = membership_fee
                    else:
                        membership_fee = int(membership_fee) + 1
                    # print(id,date_year,date_half_year,name,ID_num,salary,membership_fee,personnel_type)
                    member = Member(member_id=member_id, date_year=date_year, date_half_year=date_half_year, name=name,
                                    ID_num=ID_num, salary=salary, membership_fee=membership_fee, personnel_type=personnel_type)
                    db.session.add(member)
                db.session.commit()
                wb.close()
                os.rename("D:/工会财管/会费信息.xlsx", "D:/工会财管/会费信息导入" +
                          str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + ".xlsx")
                flash("会费信息批量导入完成")
            else:
                flash("会费文档不存在！")
            #return redirect(url_for('membership_list'))
            return render_template('membership_list.html', title='会费情况', form=form)
        if request.form['request_button'] == "导出Excel":
            if len(filters) == 0:
                flash("请输入查询条件！")
                return redirect(url_for('membership_list'))
            if members is None:
                flash('查无数据！')
                return redirect(url_for('membership_list'))
            else:
                wb = Workbook()
                sheet = wb.active
                sheet.title = '会费查询'
                sheet['A1'].value = '会员号'
                sheet['B1'].value = '姓名'
                sheet['C1'].value = '职工类型'
                sheet['D1'].value = '月工资'
                sheet['E1'].value = '半年会费'
                for member in members:
                    i = sheet.max_row + 1
                    sheet.cell(row = i, column = 1).value = member.member_id
                    sheet.cell(row = i, column = 2).value = member.name
                    sheet.cell(row = i, column = 3).value = member.personnel_type
                    sheet.cell(row = i, column = 4).value = member.salary
                    sheet.cell(row = i, column = 5).value = member.membership_fee
                #file_path = tkinter.filedialog.asksaveasfilename(filetypes=[('xlsx', '*.xlsx')],initialdir='D:/工会财管/')
                #filename = filename + '.xls'
                #sheet.column_dimensions['A'].width = 4.5
                #sheet.column_dimensions['B'].width = 12
                #sheet.column_dimensions['C'].width = 9
                #sheet.column_dimensions['D'].width = 9
                #sheet.column_dimensions['F'].width = 38
                wb.save("D:/工会财管/会费数据导出" + \
                        str(time.strftime('%Y-%m-%d %H.%M.%S')) + '.xlsx')
                wb.close()
                flash('已导出到' + "D:/工会财管/会费数据导出" + \
                        str(time.strftime('%Y-%m-%d %H.%M.%S')) + '.xlsx')
            return render_template('membership_list.html', title="会费情况",form=form,is_ok=0)
    return render_template('membership_list.html', title="会费情况",form=form,is_ok=0)


@login_required
def membership_roster():
    form = MemberInfoForm()
    if request.method == 'POST':
        filters = []
        if form.date_year.data:
            filters.append(Member.date_year == form.date_year.data)
        else:
            flash('请输入年份')
            return render_template('membership_roster.html', title='会员信息',form=form)
        if form.date_half_year.data:
            filters.append(Member.date_half_year == form.date_half_year.data)
        else:
            flash('请输入半年区间')
            return render_template('membership_roster.html', title='会员信息',form=form)
        if form.name.data:
            filters.append(Member.name == form.name.data)
        if form.personnel_type.data:
            filters.append(Member.personnel_type == form.personnel_type.data)
        if form.gender.data:
            filters.append(Member_info.gender == form.gender.data)
        if form.political_status.data:
            filters.append(Member_info.political_status ==
                           form.political_status.data)
        members = db.session.query(Member.member_id,Member.name,Member_info.gender,Member.personnel_type,Member_info.nation,Member_info.political_status,Member_info.native_place,Member_info.date_birth,Member_info.education,Member.ID_num,Member_info.phone,Member_info.email,Member_info.remark).join(Member_info,Member.ID_num == Member_info.ID_num).filter(*filters).all()
        # ---------------------------以下各功能按钮----------------------------------------------
        if request.form['request_button'] == "查询":
            return render_template('membership_roster.html', title='会员信息',form=form,members=members,year=form.date_year.data,half_year=form.date_half_year.data)
        # 导入会员信息
        # 生成互助花名册
        if request.form['request_button'] == "导入Excel":
            if os.access("D:/工会财管/会员信息.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('D:/工会财管/会员信息.xlsx')
                sheet = wb['会员信息']
                if sheet.max_row == 1:
                    flash("数据表为空表！")
                    return redirect(url_for('test'))
                for x in range(2, sheet.max_row+1):
                    id = sheet.cell(row=x, column=1).value
                    ID_num = str(sheet.cell(row=x, column=3).value)      #.strftime('%Y')
                    if sheet.cell(row=x, column=4).value is None:
                        date_birth = ''
                    else:
                        date_birth = sheet.cell(row=x, column=4).value
                    if sheet.cell(row=x, column=5).value is None:
                        gender = ''
                    else:
                        gender = sheet.cell(row=x, column=5).value
                    if sheet.cell(row=x, column=6).value is None:
                        nation = ''
                    else:
                        nation = sheet.cell(row=x, column=6).value
                    if sheet.cell(row=x, column=7).value is None:
                        political_status = ''
                    else:
                        political_status = sheet.cell(row=x, column=7).value 
                    if sheet.cell(row=x, column=8).value is None:
                        native_palce = ''
                    else:
                        native_place = sheet.cell(row=x, column=8).value
                    if sheet.cell(row=x, column=9).value is None:
                        education = ''
                    else:
                        education = sheet.cell(row=x, column=9).value
                    if sheet.cell(row=x, column=10).value is None:
                        phone = ''
                    else:
                        phone = sheet.cell(row=x, column=10).value
                    if sheet.cell(row=x, column=11).value is None:
                        email = ''
                    else:
                        email = sheet.cell(row=x, column=11).value
                    if sheet.cell(row=x, column=12).value is None:
                        remark = ''
                    else:
                        remark = sheet.cell(row=x, column=12).value
                    # print(id,ID_num,date_birth,gender,nation,political_status,native_place,education,phone,email,remark)
                    member_info = Member_info(ID_num=ID_num, date_birth=date_birth,gender=gender,nation=nation,political_status=political_status,native_place=native_place,education=education,phone=phone,email=email,remark=remark)
                    db.session.add(member_info)
                db.session.commit()
                wb.close()
                os.rename("D:/工会财管/会员信息.xlsx", "D:/工会财管/会员信息导入" + str(time.strftime("%Y-%m-%d %H.%M.%S", time.localtime())) + ".xlsx")
                flash("会员信息批量导入完成")
            else:
                flash("会员文档不存在！")
            return render_template('membership_roster.html', title="会员花名册",form=form)
        if request.form['request_button'] == "导出互助数据":
            filters = []
            if form.date_year.data:
                filters.append(Member.date_year == form.date_year.data)
            else:
                flash('请输入年份')
                return render_template('test.html', title='测试网页代码',form=form)
            if form.date_half_year.data:
                filters.append(Member.date_half_year ==
                               form.date_half_year.data)
            else:
                flash('请输入半年区间')
                return render_template('test.html', title='测试网页代码',form=form)
            members = db.session.query(Member.id,Member.name,Member_info.gender,Member.ID_num,Member_info.phone).join(Member_info,Member.ID_num == Member_info.ID_num).filter(*filters).all()
            #members = Member_info.query.order_by(Member_info.id).filter(*filters).all()
            if os.access("D:/工会财管/数据模板/机关事业-模板.xlsx", os.F_OK):
                wb = openpyxl.load_workbook('D:/工会财管/数据模板/机关事业-模板.xlsx')
                sheet = wb['事业2']
                i = 3
                for member in members:
                    i = i + 1
                    # print(i-3,member.name,member.gender,member.ID_num,member.phone)
                    sheet.cell(row = i, column = 1,value = i-3)
                    sheet.cell(row = i, column = 2,value = member.name)
                    sheet.cell(row = i, column = 3,value = member.gender)
                    sheet.cell(row = i, column = 4,value = member.ID_num)
                    sheet.cell(row = i, column = 5,value = member.phone)
                wb.save('D:/工会财管/石狮市永宁中学-机关事业-1607395530.xlsx')
                wb.close()
                flash("互助信息已导出，文件名：D:/工会财管/数据模板/石狮市永宁中学-机关事业-1607395530.xlsx")
            else:
                flash("互助模板不存在，文件名：机关事业-模板.xlsx")
            return render_template('membership_roster.html', title="会员花名册",form=form) 
    return render_template('membership_roster.html', title="会员花名册",form=form)


def buju():
    form = AccountQueryForm()
    return render_template('buju.html', title='布局',form=form)


def accounts():
    return render_template('accounts.html', title='布局2')


def page_not_found(e):
    return render_template('404.html', title="404"),404
